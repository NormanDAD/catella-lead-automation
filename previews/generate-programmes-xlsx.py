import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/sessions/tender-wonderful-goldberg/mnt/lead-automation/programmes.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

rows = sorted(data.values(), key=lambda p: p['nom_commercial'].lower())

wb = Workbook()
ws = wb.active
ws.title = "Programmes"

headers = [
    "#",
    "Nom commercial",
    "Ville",
    "CP",
    "Département",
    "Promoteur",
    "Accroche",
    "Nom du fichier PDF attendu",
    "Brochure récupérée ?",
    "Source de la brochure",
    "Notes",
]
ws.append(headers)

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', start_color='1F2937')
header_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_side = Side(border_style='thin', color='D1D5DB')
header_border = Border(bottom=Side(border_style='medium', color='111827'))

for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = header_border

body_font = Font(name='Arial', size=10)
body_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
zebra_fill = PatternFill('solid', start_color='F9FAFB')
todo_fill = PatternFill('solid', start_color='FEF3C7')

for i, prog in enumerate(rows, start=1):
    excel_row = i + 1
    nom = prog.get('nom_commercial', '')
    ws.cell(row=excel_row, column=1, value=i)
    ws.cell(row=excel_row, column=2, value=nom)
    ws.cell(row=excel_row, column=3, value=prog.get('ville', ''))
    ws.cell(row=excel_row, column=4, value=prog.get('code_postal', ''))
    ws.cell(row=excel_row, column=5, value=prog.get('departement', ''))
    ws.cell(row=excel_row, column=6, value=prog.get('promoteur', ''))
    ws.cell(row=excel_row, column=7, value=prog.get('accroche', ''))
    ws.cell(row=excel_row, column=8, value=f"{nom}.pdf")
    ws.cell(row=excel_row, column=9, value="")
    ws.cell(row=excel_row, column=10, value="")
    ws.cell(row=excel_row, column=11, value="")

    for col in range(1, 12):
        c = ws.cell(row=excel_row, column=col)
        c.font = body_font
        c.alignment = body_align
        c.border = Border(bottom=thin_side)
        if i % 2 == 0:
            c.fill = zebra_fill
        if col in (9, 10, 11):
            c.fill = todo_fill

widths = {1: 5, 2: 28, 3: 22, 4: 7, 5: 6, 6: 18, 7: 60, 8: 32, 9: 18, 10: 28, 11: 30}
for col_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 32
for i in range(2, len(rows) + 2):
    ws.row_dimensions[i].height = 48

ws.freeze_panes = "A2"
ws.sheet_view.showGridLines = False

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

legend_row = len(rows) + 3
ws.cell(row=legend_row, column=2, value="Légende :").font = Font(name='Arial', bold=True, size=10)
ws.cell(row=legend_row+1, column=2, value="Colonnes jaunes = à remplir au fur et à mesure de la récupération des brochures").font = Font(name='Arial', size=9, italic=True, color='6B7280')
ws.cell(row=legend_row+2, column=2, value="Le nom du fichier PDF doit correspondre EXACTEMENT au 'Nom commercial' pour que Power Automate l'attache au bon mail.").font = Font(name='Arial', size=9, italic=True, color='6B7280')

out = '/sessions/tender-wonderful-goldberg/mnt/lead-automation/Programmes Catella.xlsx'
wb.save(out)
print(f"OK — {len(rows)} programmes exportés → {out}")
